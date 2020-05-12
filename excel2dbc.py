#!/usr/bin/env python
# -*- coding:utf-8 -*-
# Author: 熊书灵
# CreatTime: 2020/4/18
# Email: shulingxiong@163.com


import openpyxl
import xlrd
# 导入Font对象 和 colors 颜色常量、背景颜色
from openpyxl.styles import Font, colors
from openpyxl.styles import PatternFill


class Excel2Dbc:
    def __init__(self, filepath, filename):
        self.filepath = filepath  # 文件路径
        self.filename = filename  # 无后缀的文件路径
        print(self.filename)
        self.head = ['VERSION ""\n',
                     '\n',
                     '\n',
                     'NS_ :\n',
                     '\tNS_DESC_\n', '\tCM_\n', '\tBA_DEF_\n', '\tBA_\n', '\tVAL_\n','\tCAT_DEF_\n', '\tCAT_\n',
                     '\tFILTER\n', '\tBA_DEF_DEF_\n', '\tEV_DATA_\n','\tENVVAR_DATA_\n', '\tSGTYPE_\n',
                     '\tSGTYPE_VAL_\n', '\tBA_DEF_SGTYPE_\n', '\tBA_SGTYPE_\n','\tSIG_TYPE_REF_\n', '\tVAL_TABLE_\n',
                     '\tSIG_GROUP_\n', '\tSIG_VALTYPE_\n', '\tSIGTYPE_VALTYPE_\n', '\tBO_TX_BU_\n', '\tBA_DEF_REL_\n',
                     '\tBA_REL_\n', '\tBA_DEF_DEF_REL_\n', '\tBU_SG_REL_\n', '\tBU_EV_REL_\n', '\tBU_BO_REL_\n',
                     '\tSG_MUL_VAL_\n', '\n', 'BS_:\n', '\n', 'BU_: BMS\n', '\n'
                     ]
        self.first_row = []
        self.first_col = []
        self.all_list = []  # 存储表格所有数据
        self.temp_list = []  # 存储一行数据

        self.workbook = openpyxl.load_workbook(self.filepath)  # 创建一个Excel workbook 对象,并且打开一个现有文件
        self.sheet = self.workbook.active  # 获取活跃表单
        # sheet = workbook['Matrix']  # index 或 表单名字都可以
        # print(sheet.title)
        # 打印所有表单
        # print(workbook.sheetnames)
        self.rows = 0
        self.cols = 0

        self.flag = 0
        self.i = 1
        self.text = ''  # 文本

        self.traverseExcel()

    def traverseExcel(self):

        self.rows = self.sheet.max_row  # 行数，多记一行？
        self.cols = self.sheet.max_column  # 列数，多记一列？
        # print('该excel文件有 {} 行'.format(rows))
        # print('该excel文件有 {} 列'.format(cols))

        # 遍历表格
        for row in range(1, self.rows):
            for col in range(1, self.cols):
                self.temp_list.append(self.sheet.cell(row=row, column=col).value)
            self.all_list.append(self.temp_list)

        # 取第一行，title
        for i in range(1, self.cols):
            self.first_row.append(self.sheet.cell(row=1, column=i).value)  # 元素为字符串

        # 取第一列，报文名
        for i in range(1, self.rows):
            self.first_col.append(self.sheet.cell(row=i, column=1).value)  # 元素为字符串

        self.writefile1()

    def writefile1(self):
        # BO_ 部分
        # 0 ~ 182
        for value in self.first_col:
            if self.flag != value:
                self.flag = value
                print('i = %d, val = %s' % (self.i, value))

                if self.i == 1:
                    self.i += 1
                    continue  # 第一行不处理，直接结束本次循环

                self.text += '\n'
                self.text += 'BO_ '

                j = 0
                for val in self.first_row:
                    j += 1
                    if 'Msg ID' in val:
                        try:
                            # 十六转十进制
                            self.text += str(int(self.sheet.cell(row=self.i, column=j).value, base=16)) + ' '
                            break
                        except TypeError as result:
                            print(result)
                            print("i = %d, j = %d" % (self.i, j))

                j = 0
                for val in self.first_row:
                    j += 1
                    if "Msg Name" in val:
                        self.text += str(self.sheet.cell(row=self.i, column=j).value) + ': '
                        break
                j = 0
                for val in self.first_row:
                    j += 1
                    if "Msg Length" in val:
                        self.text += str(self.sheet.cell(row=self.i, column=j).value) + ' '
                        break
                j = 0
                for val in self.first_row:
                    j += 1
                    if "Transmit Node" in val:
                        if str(self.sheet.cell(row=self.i, column=j).value) != 'BMS':
                            self.text += 'Vector__XXX\n'
                            break
                        else:
                            self.text += 'BMS\n'
                            break

                self.writefile2()

            else:
                self.writefile2()

        self.writefile3()

    def writefile2(self):
        # SG_ 部分
        self.text += ' SG_ '
        j = 0
        for val in self.first_row:
            j += 1
            if 'Signal Name' in val:
                self.text += str(self.sheet.cell(row=self.i, column=j).value) + ' : '
                break
        j = 0
        for val in self.first_row:
            j += 1
            if 'Sig MSB' in val:
                self.text += str(self.sheet.cell(row=self.i, column=j).value) + '|'
                break

        j = 0
        for val in self.first_row:
            j += 1
            if 'Signal Length' in val:
                self.text += str(self.sheet.cell(row=self.i, column=j).value) + '@0+ '
                break

        j = 0
        for val in self.first_row:
            j += 1
            if 'Resolution' in val:
                self.text += '(' + str(self.sheet.cell(row=self.i, column=j).value) + ','
                break
        j = 0
        for val in self.first_row:
            j += 1
            if 'Offset' in val:
                self.text += str(self.sheet.cell(row=self.i, column=j).value) + ') '
                break
        j = 0
        for val in self.first_row:
            j += 1
            if 'Signal Min. Value' in val:
                self.text += '[' + str(self.sheet.cell(row=self.i, column=j).value) + '|'
                break
        j = 0
        for val in self.first_row:
            j += 1
            if 'Signal Max. Value' in val:
                self.text += str(self.sheet.cell(row=self.i, column=j).value) + '] '
                break
        j = 0
        for val in self.first_row:
            j += 1
            if 'Unit' in val:
                unit = str(self.sheet.cell(row=self.i, column=j).value)
                if unit == 'None':
                    unit = ''
                    self.text += '"' + unit + '" '
                    break
                else:
                    self.text += '"' + unit + '" '
                    break
        j = 0
        for val in self.first_row:
            j += 1
            if "Transmit Node" in val:
                if str(self.sheet.cell(row=self.i, column=j).value) != 'BMS':
                    self.text += 'BMS\n'
                    break
                else:
                    self.text += 'Vector__XXX\n'

        self.i += 1

    def writefile3(self):
        # CM_ SG_ 部分
        self.text += '\n\n\n'
        # 2~183
        for index in range(2, len(self.first_col) + 1):
            self.text += 'CM_ SG_ '
            j = 0
            for val in self.first_row:
                j += 1
                if "Msg ID" in val:
                    try:
                        # 十六转十进制
                        self.text += str(int(self.sheet.cell(row=index, column=j).value, base=16)) + ' '
                        break
                    except TypeError as result:
                        print(result)
                        print("i = %d, j = %d" % (index, j))
            j = 0
            for val in self.first_row:
                j += 1
                if "Signal Name" in val:
                    self.text += str(self.sheet.cell(row=index, column=j).value) + ' '
                    break

            j = 0
            for val in self.first_row:
                j += 1
                if "Signal comment" in val:
                    comment = str(self.sheet.cell(row=index, column=j).value)
                    if comment == 'None':
                        comment = ' '
                        self.text += '"' + comment + ' ";\n'
                        break
                    else:
                        self.text += '"' + comment + ' ";\n'
                        break
        self.writefile4()

    def writefile4(self):
        # BA_ 部分
        self.text += '\n'
        self.text += 'BA_ "BusType" "CAN";\n' + 'BA_ "DBName" "BMS";\n' + 'BA_ "NmNode" BU_ BMS 0;\n' + \
            'BA_ "NmAsrCanMsgCycleOffset" BU_ BMS 0;\n' + 'BA_ "NmStationAddress" BU_ BMS 0;\n' + \
            'BA_ "NmAsrNodeIdentifier" BU_ BMS 0;\n' + 'BA_ "NodeLayerModules" BU_ BMS "CANoeILNLVector.dll";\n' + \
            'BA_ "ILUsed" BU_ BMS 1;\n' + 'BA_ "NmAsrCanMsgReducedTime" BU_ BMS 320;\n' + 'BA_ "NmAsrNode" BU_ BMS 0;\n'
        self.text += 'BA_ '

        self.writeDbc()  # 写入文件

    def writeDbc(self):
        dbc_name = self.filename + '.dbc'
        with open(dbc_name, mode='w', encoding='utf-8') as f:
            print('正在写入')
            f.writelines(self.head)
            f.writelines(self.text)
            self.text = ''  # 清空
            print('写入成功，关闭文件')
            self.workbook.close()
            f.close()


